from yahoo_finance import Share
from pprint import pprint
from openpyxl import Workbook
from openpyxl import load_workbook
from qpython import qconnection
import json
import sched, time
import datetime
import sys, traceback

#init
workbook_name = "15mins_data.xlsx"
database_name = "mins_data"
start_time = "09"
check_no = 10
time_interval = 900 #15 minutes

s = sched.scheduler(time.time, time.sleep)

#HSI-Commerce & Industry x 23
stock_CI = ['0001.HK', '0019.HK', '0027.HK', '0066.HK', '0135.HK', '0144.HK', '0151.HK', '0267.HK', '0291.HK', '0293.HK', '0322.HK', '0386.HK', '0494.HK', '0700.HK', '0762.HK', '0857.HK', '0883.HK', '0941.HK', '0992.HK', '1044.HK', '1088.HK', '1880.HK', '1928.HK', '2319.HK']
#HSI-Utilities x 4
stock_U = ['0002.HK', '0003.HK', '0006.HK', '0836.HK']
#HSI-Properties x 10
stock_P = ['0004.HK', '0012.HK', '0016.HK', '0017.HK', '0083.HK', '0101.HK', '0688.HK', '0823.HK', '1109.HK', '1113.HK']
#HSI-Finance x 12
stock_F = ['0005.HK', '0011.HK', '0023.HK', '0388.HK', '0939.HK', '1299.HK', '1398.HK', '2318.HK', '2388.HK', '2628.HK', '3328.HK', '3988.HK']

temp = {} #for checking duplicate data
for stock in stock_CI:
	temp[stock] = 0
for stock in stock_U:
	temp[stock] = 0
for stock in stock_P:
	temp[stock] = 0
for stock in stock_F:
	temp[stock] = 0
count = 0 #for counting collected data
duplicate = 0 #check for valid append
overall_duplicate = 0 #check for end of date
s_CI = {}
s_U = {}
s_P = {}
s_F = {}
ws_CI = {}
ws_U = {}
ws_P = {}
ws_F = {}

def define_stock():
	print("================================== Define Stock ===================================")
	#HSI-Commerce & Industry
	for stock in stock_CI:
		while True:
			try:
				s_CI[stock] = Share(stock)
			except:
				print("Failed loading for " + stock + "! Try again...")
				continue
			break
		print("Finished " + stock)

	#HSI-Utilities
	for stock in stock_U:
		while True:
			try:
				s_U[stock] = Share(stock)
			except:
				print("Failed loading for " + stock + "! Try again...")
				continue
			break
		print("Finished " + stock)
	#HSI-Properties
	for stock in stock_P:
		while True:
			try:
				s_P[stock] = Share(stock)
			except:
				print("Failed loading for " + stock + "! Try again...")
				continue
			break
		print("Finished " + stock)
	#HSI-Finance
	for stock in stock_F:
		while True:
			try:
				s_F[stock] = Share(stock)
			except:
				print("Failed loading for " + stock + "! Try again...")
				continue
			break
		print("Finished " + stock)

def refresh_stock():
	print("================================== Stock Refresh ==================================")
	#HSI-Commerce & Industry
	for stock in stock_CI:
		s_CI[stock].refresh
	#HSI-Utilities
	for stock in stock_U:
		s_U[stock].refresh
	#HSI-Properties
	for stock in stock_P:
		s_P[stock].refresh
	#HSI-Finance
	for stock in stock_F:
		s_F[stock].refresh

def connect_db():
	q = qconnection.QConnection(host='localhost', port=5000)
	q.open()
	print("================================ Connection Details ===============================")
	print(q)
	print('IPC version: %s. Is connected: %s' % (q.protocol_version, q.is_connected()))
	return q;

def write_workbook():
	q = connect_db()
	print("================================== Write Workbook =================================")
	global duplicate
	wb = load_workbook(workbook_name)
	#HSI-Commerce & Industry
	for stock in stock_CI:
		try:
			ws_CI[stock] = wb.get_sheet_by_name(stock)
		except:
			ws_CI[stock] = wb.create_sheet(title=stock)
			ws_CI[stock].append(["Date", "Price", "Change", "Volume", "Open", "Avg Daily Volume", "Stock exchange", "Market Cap", "Book Value", "Ebitda", "Dividend share", "Dividend yield", "Earnings Share", "Days High", "Days Low", "50day moving avg", "200day moving avg", "Price earnings ratio", "Price earnings growth ratio", "Price sales", "Price book", "Short Ratio"])
		while True:
			try:
				temp_val = s_CI[stock].get_trade_datetime()
				if temp_val == temp[stock]:
					print("Duplicate data detected for " + stock)
					duplicate = duplicate + 1
					break

				temp_array = [s_CI[stock].get_trade_datetime(), s_CI[stock].get_price(), s_CI[stock].get_change(), s_CI[stock].get_volume(), s_CI[stock].get_open(), s_CI[stock].get_avg_daily_volume(), s_CI[stock].get_stock_exchange(), s_CI[stock].get_market_cap(), s_CI[stock].get_book_value(), s_CI[stock].get_ebitda(), s_CI[stock].get_dividend_share(), s_CI[stock].get_dividend_yield(), s_CI[stock].get_earnings_share(), s_CI[stock].get_days_high(), s_CI[stock].get_days_low(), s_CI[stock].get_50day_moving_avg(), s_CI[stock].get_200day_moving_avg(), s_CI[stock].get_price_earnings_ratio(), s_CI[stock].get_price_earnings_growth_ratio(), s_CI[stock].get_price_sales(), s_CI[stock].get_price_book(), s_CI[stock].get_short_ratio()]
				
				ws_CI[stock].append(temp_array)

				temp_array_count = 0
				for attr in temp_array:
					if not attr:
						temp_array[temp_array_count] = '999.9'
					temp_array_count += 1

				if temp_array[2][0] == '+':
					temp_array[2] = temp_array[2].replace("+", "u")
				elif temp_array[2][0] == '-':
					temp_array[2] = temp_array[2].replace("-", "d")

				query = '`' + str(database_name) + ' insert(`' + str(stock) + ';' + str(temp_array[0][11:19]) + ';' + str(temp_array[0][0:10].replace("-", ".")) + ';' + str(float(temp_array[1])) + ';`' + str(temp_array[2]) + ';' + str(float(temp_array[3])) + ';' + str(float(temp_array[4])) + ';' + str(float(temp_array[5])) + ';`' + str(temp_array[6]) + ';`' + str(temp_array[7]) + ';' + str(float(temp_array[8])) + ';`' + str(temp_array[9]) + ';' + str(float(temp_array[10])) + ';' + str(float(temp_array[11])) + ';' + str(float(temp_array[12])) + ';' + str(float(temp_array[13])) + ';' + str(float(temp_array[14])) + ';' + str(float(temp_array[15])) + ';' + str(float(temp_array[16])) + ';' + str(float(temp_array[17])) + ';' + str(float(temp_array[18])) + ';' + str(float(temp_array[19])) + ';' + str(float(temp_array[20])) + ';' + str(float(temp_array[21])) + ')'
				while True:
					try:
						df = q(query)
					except:
						print("Failed inserting for " + str(sheet.title) + "! Try again...")
						continue
					break
				temp[stock] = temp_val
			except:
				print("Failed writing for " + stock + "! Try again...")
				continue
			break
	print("Wrote workbook for stock_CI")
	#HSI-Utilities
	for stock in stock_U:
		try:
			ws_U[stock] = wb.get_sheet_by_name(stock)
		except:
			ws_U[stock] = wb.create_sheet(title=stock)
			ws_U[stock].append(["Date", "Price", "Change", "Volume", "Open", "Avg Daily Volume", "Stock exchange", "Market Cap", "Book Value", "Ebitda", "Dividend share", "Dividend yield", "Earnings Share", "Days High", "Days Low", "50day moving avg", "200day moving avg", "Price earnings ratio", "Price earnings growth ratio", "Price sales", "Price book", "Short Ratio"])
		while True:
			try:
				temp_val = s_U[stock].get_trade_datetime()
				if temp_val == temp[stock]:
					print("Duplicate data detected for " + stock)
					duplicate = duplicate + 1
					break

				temp_array = [s_U[stock].get_trade_datetime(), s_U[stock].get_price(), s_U[stock].get_change(), s_U[stock].get_volume(), s_U[stock].get_open(), s_U[stock].get_avg_daily_volume(), s_U[stock].get_stock_exchange(), s_U[stock].get_market_cap(), s_U[stock].get_book_value(), s_U[stock].get_ebitda(), s_U[stock].get_dividend_share(), s_U[stock].get_dividend_yield(), s_U[stock].get_earnings_share(), s_U[stock].get_days_high(), s_U[stock].get_days_low(), s_U[stock].get_50day_moving_avg(), s_U[stock].get_200day_moving_avg(), s_U[stock].get_price_earnings_ratio(), s_U[stock].get_price_earnings_growth_ratio(), s_U[stock].get_price_sales(), s_U[stock].get_price_book(), s_U[stock].get_short_ratio()]
				
				ws_U[stock].append(temp_array)

				temp_array_count = 0
				for attr in temp_array:
					if not attr:
						temp_array[temp_array_count] = '999.9'
					temp_array_count += 1

				if temp_array[2][0] == '+':
					temp_array[2] = temp_array[2].replace("+", "u")
				elif temp_array[2][0] == '-':
					temp_array[2] = temp_array[2].replace("-", "d")

				query = '`' + str(database_name) + ' insert(`' + str(stock) + ';' + str(temp_array[0][11:19]) + ';' + str(temp_array[0][0:10].replace("-", ".")) + ';' + str(float(temp_array[1])) + ';`' + str(temp_array[2]) + ';' + str(float(temp_array[3])) + ';' + str(float(temp_array[4])) + ';' + str(float(temp_array[5])) + ';`' + str(temp_array[6]) + ';`' + str(temp_array[7]) + ';' + str(float(temp_array[8])) + ';`' + str(temp_array[9]) + ';' + str(float(temp_array[10])) + ';' + str(float(temp_array[11])) + ';' + str(float(temp_array[12])) + ';' + str(float(temp_array[13])) + ';' + str(float(temp_array[14])) + ';' + str(float(temp_array[15])) + ';' + str(float(temp_array[16])) + ';' + str(float(temp_array[17])) + ';' + str(float(temp_array[18])) + ';' + str(float(temp_array[19])) + ';' + str(float(temp_array[20])) + ';' + str(float(temp_array[21])) + ')'
				while True:
					try:
						df = q(query)
					except:
						print("Failed inserting for " + str(sheet.title) + "! Try again...")
						continue
					break
				temp[stock] = temp_val
			except:
				print("Failed writing for " + stock + "! Try again...")
				continue
			break
	print("Wrote workbook for stock_U")
	#HSI-Properties
	for stock in stock_P:
		try:
			ws_P[stock] = wb.get_sheet_by_name(stock)
		except:
			ws_P[stock] = wb.create_sheet(title=stock)
			ws_P[stock].append(["Date", "Price", "Change", "Volume", "Open", "Avg Daily Volume", "Stock exchange", "Market Cap", "Book Value", "Ebitda", "Dividend share", "Dividend yield", "Earnings Share", "Days High", "Days Low", "50day moving avg", "200day moving avg", "Price earnings ratio", "Price earnings growth ratio", "Price sales", "Price book", "Short Ratio"])
		while True:
			try:
				temp_val = s_P[stock].get_trade_datetime()
				if temp_val == temp[stock]:
					print("Duplicate data detected for " + stock)
					duplicate = duplicate + 1
					break

				temp_array = [s_P[stock].get_trade_datetime(), s_P[stock].get_price(), s_P[stock].get_change(), s_P[stock].get_volume(), s_P[stock].get_open(), s_P[stock].get_avg_daily_volume(), s_P[stock].get_stock_exchange(), s_P[stock].get_market_cap(), s_P[stock].get_book_value(), s_P[stock].get_ebitda(), s_P[stock].get_dividend_share(), s_P[stock].get_dividend_yield(), s_P[stock].get_earnings_share(), s_P[stock].get_days_high(), s_P[stock].get_days_low(), s_P[stock].get_50day_moving_avg(), s_P[stock].get_200day_moving_avg(), s_P[stock].get_price_earnings_ratio(), s_P[stock].get_price_earnings_growth_ratio(), s_P[stock].get_price_sales(), s_P[stock].get_price_book(), s_P[stock].get_short_ratio()]
				
				ws_P[stock].append(temp_array)

				temp_array_count = 0
				for attr in temp_array:
					if not attr:
						temp_array[temp_array_count] = '999.9'
					temp_array_count += 1

				if temp_array[2][0] == '+':
					temp_array[2] = temp_array[2].replace("+", "u")
				elif temp_array[2][0] == '-':
					temp_array[2] = temp_array[2].replace("-", "d")

				query = '`' + str(database_name) + ' insert(`' + str(stock) + ';' + str(temp_array[0][11:19]) + ';' + str(temp_array[0][0:10].replace("-", ".")) + ';' + str(float(temp_array[1])) + ';`' + str(temp_array[2]) + ';' + str(float(temp_array[3])) + ';' + str(float(temp_array[4])) + ';' + str(float(temp_array[5])) + ';`' + str(temp_array[6]) + ';`' + str(temp_array[7]) + ';' + str(float(temp_array[8])) + ';`' + str(temp_array[9]) + ';' + str(float(temp_array[10])) + ';' + str(float(temp_array[11])) + ';' + str(float(temp_array[12])) + ';' + str(float(temp_array[13])) + ';' + str(float(temp_array[14])) + ';' + str(float(temp_array[15])) + ';' + str(float(temp_array[16])) + ';' + str(float(temp_array[17])) + ';' + str(float(temp_array[18])) + ';' + str(float(temp_array[19])) + ';' + str(float(temp_array[20])) + ';' + str(float(temp_array[21])) + ')'
				while True:
					try:
						df = q(query)
					except:
						print("Failed inserting for " + str(sheet.title) + "! Try again...")
						continue
					break
				temp[stock] = temp_val
			except:
				print("Failed writing for " + stock + "! Try again...")
				continue
			break
	print("Wrote workbook for stock_P")
	#HSI-Finance
	for stock in stock_F:
		try:
			ws_F[stock] = wb.get_sheet_by_name(stock)
		except:
			ws_F[stock] = wb.create_sheet(title=stock)
			ws_F[stock].append(["Date", "Price", "Change", "Volume", "Open", "Avg Daily Volume", "Stock exchange", "Market Cap", "Book Value", "Ebitda", "Dividend share", "Dividend yield", "Earnings Share", "Days High", "Days Low", "50day moving avg", "200day moving avg", "Price earnings ratio", "Price earnings growth ratio", "Price sales", "Price book", "Short Ratio"])
		while True:
			try:
				temp_val = s_F[stock].get_trade_datetime()
				if temp_val == temp[stock]:
					print("Duplicate data detected for " + stock)
					duplicate = duplicate + 1
					break

				temp_array = [s_F[stock].get_trade_datetime(), s_F[stock].get_price(), s_F[stock].get_change(), s_F[stock].get_volume(), s_F[stock].get_open(), s_F[stock].get_avg_daily_volume(), s_F[stock].get_stock_exchange(), s_F[stock].get_market_cap(), s_F[stock].get_book_value(), s_F[stock].get_ebitda(), s_F[stock].get_dividend_share(), s_F[stock].get_dividend_yield(), s_F[stock].get_earnings_share(), s_F[stock].get_days_high(), s_F[stock].get_days_low(), s_F[stock].get_50day_moving_avg(), s_F[stock].get_200day_moving_avg(), s_F[stock].get_price_earnings_ratio(), s_F[stock].get_price_earnings_growth_ratio(), s_F[stock].get_price_sales(), s_F[stock].get_price_book(), s_F[stock].get_short_ratio()]
				
				ws_F[stock].append(temp_array)

				temp_array_count = 0
				for attr in temp_array:
					if not attr:
						temp_array[temp_array_count] = '999.9'
					temp_array_count += 1

				if temp_array[2][0] == '+':
					temp_array[2] = temp_array[2].replace("+", "u")
				elif temp_array[2][0] == '-':
					temp_array[2] = temp_array[2].replace("-", "d")

				query = '`' + str(database_name) + ' insert(`' + str(stock) + ';' + str(temp_array[0][11:19]) + ';' + str(temp_array[0][0:10].replace("-", ".")) + ';' + str(float(temp_array[1])) + ';`' + str(temp_array[2]) + ';' + str(float(temp_array[3])) + ';' + str(float(temp_array[4])) + ';' + str(float(temp_array[5])) + ';`' + str(temp_array[6]) + ';`' + str(temp_array[7]) + ';' + str(float(temp_array[8])) + ';`' + str(temp_array[9]) + ';' + str(float(temp_array[10])) + ';' + str(float(temp_array[11])) + ';' + str(float(temp_array[12])) + ';' + str(float(temp_array[13])) + ';' + str(float(temp_array[14])) + ';' + str(float(temp_array[15])) + ';' + str(float(temp_array[16])) + ';' + str(float(temp_array[17])) + ';' + str(float(temp_array[18])) + ';' + str(float(temp_array[19])) + ';' + str(float(temp_array[20])) + ';' + str(float(temp_array[21])) + ')'
				while True:
					try:
						df = q(query)
					except:
						print("Failed inserting for " + str(sheet.title) + "! Try again...")
						continue
					break
				temp[stock] = temp_val
			except:
				print("Failed writing for " + stock + "! Try again...")
				continue
			break
	print("Wrote workbook for stock_F")

	wb.save(workbook_name)

def check():
	global duplicate
	global overall_duplicate
	global count
	if duplicate < 50:
		count = count + 1
		duplicate = 0
	elif duplicate == 50:
		overall_duplicate = overall_duplicate + 1
		duplicate = 0
	else:
		print("Error in Checking!")
	print("\nTotal data collected: " + str(count))
	print("Total overall_duplicate: " + str(overall_duplicate))
	return overall_duplicate;

def main(sc):
	global overall_duplicate
	while True:
		try:
			define_stock()
			refresh_stock()
			write_workbook()
		except:
			define_stock()
			continue
		break
	if check() == check_no:
		overall_duplicate = 0
		print('program restarting at: ' + str(datetime.datetime.now().time()))
		while str(datetime.datetime.now().time())[0:2] != start_time:
			if(str(datetime.datetime.now().time())[3:5] == str(datetime.datetime.now().time())[6:8] == '00' and str(datetime.datetime.now().time())[9:14] == '00000'):
				print(datetime.datetime.now().time())
		sc.enter(1, 1, main, (sc, ))
	else:
		sc.enter(time_interval, 1, main, (sc, ))

print('program start at: ' + str(datetime.datetime.now().time()))
while str(datetime.datetime.now().time())[0:2] != start_time:
	if(str(datetime.datetime.now().time())[3:5] == str(datetime.datetime.now().time())[6:8] == '00' and str(datetime.datetime.now().time())[9:14] == '00000'):
		print(datetime.datetime.now().time())

s.enter(1, 1, main, (s, ))
s.run()