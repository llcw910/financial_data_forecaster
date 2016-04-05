from yahoo_finance import Share
from pprint import pprint
from openpyxl import Workbook
from openpyxl import load_workbook
from qpython import qconnection
import sched, time
import datetime
import sys, traceback
import os
import time

#init
old_workbook_name = "stock_update.xlsx"
workbook_name = "15mins_data.xlsx"
database_name = "mins_data"
attr_array = ["Date", "Price", "Change", "Volume", "Open", "Avg Daily Volume", "Stock exchange", "Market Cap", "Book Value", "Ebitda", "Dividend share", "Dividend yield", "Earnings Share", "Days High", "Days Low", "50day moving avg", "200day moving avg", "Price earnings ratio", "Price earnings growth ratio", "Price sales", "Price book", "Short Ratio"]

#HSI-Commerce & Industry x 24
stock_CI = ['0001.HK', '0019.HK', '0027.HK', '0066.HK', '0135.HK', '0144.HK', '0151.HK', '0267.HK', '0291.HK', '0293.HK', '0322.HK', '0386.HK', '0494.HK', '0700.HK', '0762.HK', '0857.HK', '0883.HK', '0941.HK', '0992.HK', '1044.HK', '1088.HK', '1880.HK', '1928.HK', '2319.HK']
#HSI-Utilities x 4
stock_U = ['0002.HK', '0003.HK', '0006.HK', '0836.HK']
#HSI-Properties x 10
stock_P = ['0004.HK', '0012.HK', '0016.HK', '0017.HK', '0083.HK', '0101.HK', '0688.HK', '0823.HK', '1109.HK', '1113.HK']
#HSI-Finance x 12
stock_F = ['0005.HK', '0011.HK', '0023.HK', '0388.HK', '0939.HK', '1299.HK', '1398.HK', '2318.HK', '2388.HK', '2628.HK', '3328.HK', '3988.HK']

def connect_db():
	q = qconnection.QConnection(host='localhost', port=5000)
	q.open()
	print("================================= Connection Details ==================================")
	print(q)
	print('IPC version: %s. Is connected: %s' % (q.protocol_version, q.is_connected()))
	return q;

def store_db(wb_name, q):
	wb = load_workbook(wb_name)
	print("================================ Load data from excel =================================")
	for sheet in wb:
		data_count = 0;
		last_value = '-1';
		for row in sheet.iter_rows("A2:V" + str(sheet.max_row)):
			temp = {}
			count = 0;
			for cell in row:
				temp[attr_array[count]] = cell.value
				count+=1
			if temp['Date'] == last_value:
				print("Duplicated value detected for Date: " + str(temp['Date']) + "! Skipped!")
			else:
				temp_array_count = 0
				for attr in attr_array:
					if not temp[attr]:
						temp[attr] = '999.9'
					temp_array_count += 1

				if temp[attr_array[2]][0] == '+':
					temp[attr_array[2]] = temp[attr_array[2]].replace("+", "u")
				elif temp[attr_array[2]][0] == '-':
					temp[attr_array[2]] = temp[attr_array[2]].replace("-", "d")

				query = '`' + str(database_name) + ' insert(`' + str(sheet.title) + ';' + str(temp[attr_array[0]][11:19]) + ';' + str(temp[attr_array[0]][0:10].replace("-", ".")) + ';' + str(float(temp[attr_array[1]])) + ';`' + str(temp[attr_array[2]]) + ';' + str(float(temp[attr_array[3]])) + ';' + str(float(temp[attr_array[4]])) + ';' + str(float(temp[attr_array[5]])) + ';`' + str(temp[attr_array[6]]) + ';`' + str(temp[attr_array[7]]) + ';' + str(float(temp[attr_array[8]])) + ';`' + str(temp[attr_array[9]]) + ';' + str(float(temp[attr_array[10]])) + ';' + str(float(temp[attr_array[11]])) + ';' + str(float(temp[attr_array[12]])) + ';' + str(float(temp[attr_array[13]])) + ';' + str(float(temp[attr_array[14]])) + ';' + str(float(temp[attr_array[15]])) + ';' + str(float(temp[attr_array[16]])) + ';' + str(float(temp[attr_array[17]])) + ';' + str(float(temp[attr_array[18]])) + ';' + str(float(temp[attr_array[19]])) + ';' + str(float(temp[attr_array[20]])) + ';' + str(float(temp[attr_array[21]])) + ')'
				last_value = str(temp['Date'])
				while True:
					try:
						df = q(query)
						data_count += 1;
					except:
						print("Failed writing for " + str(sheet.title) + "! Try again...")
						continue
					break
		print(str(data_count) + str(df) + ' data for ' + str(sheet.title) + ' recorded.')

#================================== body ==================================
q = connect_db()
store_db(old_workbook_name, q)
store_db(workbook_name, q)